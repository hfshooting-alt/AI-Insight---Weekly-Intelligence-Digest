#!/usr/bin/env python3
"""
Daily paper scouting + deep-reading report agent.
"""

from __future__ import annotations

import datetime as dt
import html
import json
import math
import os
import pathlib
import re
import smtplib
import tempfile
import textwrap
import time
from dataclasses import dataclass
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Dict, Iterable, List, Optional, Tuple

import feedparser
import requests
from apscheduler.schedulers.blocking import BlockingScheduler
from openai import OpenAI

GEMINI_BASE_URL = "https://generativelanguage.googleapis.com/v1beta/openai/"
DEFAULT_MODEL = "gemini-3.0-flash-preview"

BEIJING_TZ = dt.timezone(dt.timedelta(hours=8))

SEARCH_TERMS = [
    "physical ai data infrastructure",
    "embodied ai data pipeline",
    "robotics data engine",
    "sim2real data pipeline",
    "autonomous driving data infrastructure",
    "3d perception data pipeline",
    "lidar data pipeline",
    "sensor fusion dataset pipeline",
    "world model for robotics",
    "robot simulation data",
]

TOPIC_KEYWORDS = SEARCH_TERMS + [
    "具身智能数据基础设施",
    "机器人数据管道",
    "自动驾驶数据基础设施",
    "3d感知数据",
    "激光雷达数据",
    "传感器融合数据",
    "仿真到现实数据",
    "世界模型",
]

WORLD_STRICT_KEYWORDS = [
    "world engine", "world model", "world simulator", "interactive world",
    "physics simulator", "robot simulation", "digital twin", "embodied",
    "世界引擎", "世界模型", "仿真引擎",
]

INFRA_STRICT_KEYWORDS = [
    "robot data pipeline", "embodied data pipeline", "autonomy data engine", "sim2real data",
    "teleoperation data", "fleet data", "sensor data pipeline", "perception data curation",
    "multimodal robot dataset", "policy dataset", "trajectory dataset", "feature store for robotics",
    "lakehouse for autonomy", "data labeling for perception", "data governance for robotics",
    "physical ai data", "具身智能数据", "机器人数据管道", "自动驾驶数据管道", "感知数据治理",
    "仿真到现实数据", "遥操作数据", "轨迹数据集", "策略数据集", "车队数据引擎",
    # Data-centric approaches for embodied/physical AI
    "data scaling law for robot", "robot data flywheel", "demonstration collection",
    "cross-embodiment data", "open x-embodiment", "droid dataset", "bridge dataset",
    "data augmentation for manipulation", "embodied data generation", "synthetic data for robot",
    "数据驱动具身", "机器人数据增强", "跨实体数据", "合成数据用于机器人",
]

IRRELEVANT_HINT_KEYWORDS = [
    "t cell", "tumor", "cancer", "clinical", "patient", "medical", "medicine",
    "drug", "antigen", "genome", "protein", "cell", "herb", "medicinal plant",
    "botany", "agriculture", "crop", "ethnobotany", "pharmacology",
    "spacecraft", "satellite", "astronomy", "orbit",
    "航天", "医学", "肿瘤", "临床", "蛋白", "细胞", "草药", "药用植物", "农业", "作物",
]

INVESTMENT_TECH_KEYWORDS = [
    "model", "engine", "simulator", "training", "inference", "benchmark", "algorithm",
    "system", "framework", "platform", "pipeline", "infrastructure", "scalable",
    "latency", "throughput", "reliability", "dataset", "evaluation", "deployment",
    "architecture", "distributed", "stream", "batch", "etl", "lakehouse",
    "模型", "引擎", "仿真", "训练", "推理", "系统", "框架", "平台", "管道", "基础设施",
]

EXCLUDED_NON_TECH_KEYWORDS = [
    "art", "artist", "mural", "museum", "heritage", "aesthetics", "culture",
    "filipino", "humanities", "literature", "music", "dance", "painting", "exhibition",
    "艺术", "壁画", "博物馆", "文化", "文学", "音乐", "舞蹈", "展览", "美学",
]

CORE_WORLD_ANCHORS = [
    "world model", "world engine", "world simulator", "interactive world", "digital twin", "世界模型", "世界引擎",
]

CORE_INFRA_ANCHORS = [
    "data infrastructure", "data infra", "data pipeline", "data ingestion", "data lakehouse", "feature store",
    "stream processing", "batch processing", "etl", "data orchestration", "数据基础设施", "数据管道", "湖仓",
    "robot data", "embodied data", "perception data", "sim2real data", "trajectory data",
    "机器人数据", "具身数据", "感知数据", "仿真数据", "轨迹数据",
]

PHYSICAL_AI_MUST_HAVE = [
    "physical ai", "embodied", "embodied ai", "robot", "robotics", "autonomous driving", "self-driving",
    "sim2real", "具身", "具身智能", "机器人", "自动驾驶", "物理智能",
]

DATA_INFRA_MUST_HAVE = [
    "data infra", "data infrastructure", "data pipeline", "data ingestion", "feature store", "lakehouse",
    "dataset", "data engine", "数据基础设施", "数据管道", "数据引擎", "数据集", "湖仓",
    # Physical AI / embodied data problems
    "robot dataset", "embodied dataset", "perception dataset", "manipulation dataset",
    "navigation dataset", "driving dataset", "sim2real dataset", "policy data",
    "trajectory data", "teleoperation data", "demonstration data", "sensor data",
    "data curation for robot", "data scaling for embodied", "data flywheel",
    "机器人数据集", "具身数据集", "感知数据集", "操作数据集", "导航数据集",
    "驾驶数据集", "策略数据", "轨迹数据", "遥操作数据", "演示数据", "数据飞轮",
]

PHYSICAL_AI_CONTEXT_KEYWORDS = [
    "embodied", "robot", "robotics", "autonomous driving", "self-driving", "carla", "airsim", "gazebo",
    "sim2real", "policy learning", "navigation", "manipulation", "perception", "lidar", "camera rig",
    "world model", "world simulator", "digital twin", "具身", "机器人", "自动驾驶", "导航", "操作", "感知", "仿真",
]


FOCUS_BUSINESS_KEYWORDS = [
    "3d perception", "3d sensing", "3d reconstruction", "3d modeling", "3d simulation", "point cloud", "lidar",
    "depth estimation", "slam", "neural rendering", "gaussian splatting", "occupancy", "bev", "sensor fusion",
    "physical ai", "embodied ai", "embodied intelligence", "robot learning", "sim2real", "autonomous driving",
    "3d感知", "三维感知", "三维建模", "3d建模", "点云", "激光雷达", "传感器融合", "物理智能", "具身智能",
    "机器人学习", "自动驾驶", "仿真", "数字孪生",
]


STRONG_FOCUS_KEYWORDS = [
    "3d perception", "3d sensing", "3d reconstruction", "3d modeling", "point cloud", "lidar", "sensor fusion",
    "slam", "sim2real", "physical ai", "embodied ai", "robotics", "autonomous driving", "self-driving",
    "3d感知", "三维感知", "三维建模", "点云", "激光雷达", "传感器融合", "具身智能", "物理智能", "自动驾驶", "机器人",
]

WEAK_DOMAIN_EXCLUDE_KEYWORDS = [
    "condition monitoring", "mechanical fault", "impeller", "bearing fault", "corrosion", "pump failure",
    "predictive maintenance", "iot-enabled", "叶轮", "腐蚀", "机油", "盐水", "故障诊断",
]

RSS_SOURCES = {
    "arXiv cs.RO": "http://export.arxiv.org/rss/cs.RO",
    "arXiv cs.CV": "http://export.arxiv.org/rss/cs.CV",
    "arXiv cs.AI": "http://export.arxiv.org/rss/cs.AI",
    "OpenReview Robotics": "https://openreview.net/group?id=roboticsfoundation.org/feed",
}

REQUEST_HEADERS = {
    "User-Agent": "daily-paper-agent/1.0 (academic-digest; mailto:report@example.com)",
}


@dataclass
class Paper:
    title: str
    url: str
    abstract: str
    source: str
    published: dt.datetime
    authors: List[str]
    institutions: List[str]
    author_orgs: List[str]
    citation_count: int = 0
    influence_score: float = 0.0


@dataclass
class AnalyzedPaper:
    paper: Paper
    category: str
    analysis_lines: List[str]
    early_score: int = 0
    discussion_score: float = 0.0


def now_utc() -> dt.datetime:
    return dt.datetime.now(dt.timezone.utc)


def now_beijing() -> dt.datetime:
    return dt.datetime.now(BEIJING_TZ)


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").lower()).strip()




def clean_org_name(name: str) -> str:
    text = re.sub(r"\s+", " ", (name or "").strip())
    text = re.sub(r"^[,;:\-]+|[,;:\-]+$", "", text)
    return text


def build_author_orgs(author_pairs: List[Tuple[str, str]]) -> List[str]:
    out: List[str] = []
    seen = set()
    for author, org in author_pairs:
        author_name = re.sub(r"\s+", " ", (author or "").strip())
        org_name = clean_org_name(org)
        if not author_name:
            continue
        label = f"{author_name}（{org_name}）" if org_name else author_name
        key = normalize(label)
        if key and key not in seen:
            seen.add(key)
            out.append(label)
    return out[:8]

def parse_iso_datetime(value: Optional[str]) -> Optional[dt.datetime]:
    if not value:
        return None
    try:
        parsed = dt.datetime.fromisoformat(value.replace("Z", "+00:00"))
        if not parsed.tzinfo:
            parsed = parsed.replace(tzinfo=dt.timezone.utc)
        return parsed.astimezone(dt.timezone.utc)
    except Exception:
        return None


def parse_struct_time(value: Optional[time.struct_time]) -> Optional[dt.datetime]:
    if not value:
        return None
    try:
        import calendar
        return dt.datetime.fromtimestamp(calendar.timegm(value), tz=dt.timezone.utc)
    except Exception:
        return None


def parse_date_parts(parts: List[int]) -> Optional[dt.datetime]:
    if not parts or not parts[0]:
        return None
    y = parts[0]
    m = parts[1] if len(parts) > 1 else 1
    d = parts[2] if len(parts) > 2 else 1
    try:
        return dt.datetime(y, m, d, tzinfo=dt.timezone.utc)
    except Exception:
        return None


def parse_date_string(value: Optional[str]) -> Optional[dt.datetime]:
    if not value:
        return None
    try:
        d = dt.date.fromisoformat(value)
        return dt.datetime(d.year, d.month, d.day, tzinfo=dt.timezone.utc)
    except Exception:
        return None


def html_strip(value: str) -> str:
    v = html.unescape(value or "")
    return re.sub(r"<[^>]+>", " ", v)


def sanitize_text(value: str, max_len: int = 0) -> str:
    txt = html_strip(value or "")
    txt = re.sub(r"\s+", " ", txt).strip()
    if max_len and len(txt) > max_len:
        txt = txt[:max_len].rstrip()
    return txt


def target_beijing_date_window() -> Tuple[dt.date, dt.date]:
    """Return inclusive window [t-17, t-11] in Beijing date."""
    today = now_beijing().date()
    start = today - dt.timedelta(days=17)
    end = today - dt.timedelta(days=11)
    return start, end


def in_target_beijing_window(published: Optional[dt.datetime]) -> bool:
    if not published:
        return False
    bj_date = published.astimezone(BEIJING_TZ).date()
    start, end = target_beijing_date_window()
    return start <= bj_date <= end


def beijing_day_window() -> Tuple[str, str]:
    """Return t-17/t-11 in Beijing as YYYY-MM-DD (for source-side filters)."""
    start, end = target_beijing_date_window()
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def topical_score(title: str, abstract: str) -> int:
    hay = normalize(f"{title} {abstract}")
    base = sum(1 for kw in TOPIC_KEYWORDS if normalize(kw) in hay)
    infra_hit = sum(1 for kw in (INFRA_STRICT_KEYWORDS + CORE_INFRA_ANCHORS + DATA_INFRA_MUST_HAVE) if normalize(kw) in hay)
    physical_hit = sum(1 for kw in (PHYSICAL_AI_CONTEXT_KEYWORDS + PHYSICAL_AI_MUST_HAVE) if normalize(kw) in hay)
    focus_hit = sum(1 for kw in FOCUS_BUSINESS_KEYWORDS if normalize(kw) in hay)
    penalty = sum(1 for kw in IRRELEVANT_HINT_KEYWORDS if normalize(kw) in hay)
    exclude = sum(1 for kw in EXCLUDED_NON_TECH_KEYWORDS if normalize(kw) in hay)
    miss_penalty = 8 if not (physical_hit >= 1 and infra_hit >= 1) else 0
    return base + infra_hit * 4 + physical_hit * 4 + focus_hit * 3 - penalty * 3 - exclude * 4 - miss_penalty


def is_physical_ai_data_infra_focus(text: str) -> bool:
    hay = normalize(text or "")
    physical_hit = sum(1 for kw in (PHYSICAL_AI_MUST_HAVE + PHYSICAL_AI_CONTEXT_KEYWORDS) if normalize(kw) in hay)
    infra_hit = sum(1 for kw in (DATA_INFRA_MUST_HAVE + CORE_INFRA_ANCHORS + INFRA_STRICT_KEYWORDS) if normalize(kw) in hay)
    focus_hit = sum(1 for kw in FOCUS_BUSINESS_KEYWORDS if normalize(kw) in hay)
    return physical_hit >= 1 and infra_hit >= 1 and (focus_hit >= 1 or infra_hit >= 2)


def is_domain_relevant(title: str, abstract: str) -> bool:
    """Strictly keep only Physical AI / Embodied AI data-infra papers."""
    hay = normalize(f"{title} {abstract}")
    if any(normalize(kw) in hay for kw in EXCLUDED_NON_TECH_KEYWORDS):
        return False
    if sum(1 for kw in IRRELEVANT_HINT_KEYWORDS if normalize(kw) in hay) >= 1:
        return False
    if sum(1 for kw in WEAK_DOMAIN_EXCLUDE_KEYWORDS if normalize(kw) in hay) >= 1:
        return False
    return is_physical_ai_data_infra_focus(hay)


def is_domain_relevant_soft(title: str, abstract: str) -> bool:
    hay = normalize(f"{title} {abstract}")
    if any(normalize(kw) in hay for kw in EXCLUDED_NON_TECH_KEYWORDS):
        return False
    if sum(1 for kw in IRRELEVANT_HINT_KEYWORDS if normalize(kw) in hay) >= 1:
        return False
    physical_hit = sum(1 for kw in (PHYSICAL_AI_MUST_HAVE + PHYSICAL_AI_CONTEXT_KEYWORDS) if normalize(kw) in hay)
    infra_hit = sum(1 for kw in (DATA_INFRA_MUST_HAVE + CORE_INFRA_ANCHORS + INFRA_STRICT_KEYWORDS) if normalize(kw) in hay)
    return physical_hit >= 1 and infra_hit >= 1


def fetch_arxiv() -> List[Paper]:
    query = " OR ".join([f'all:"{term}"' for term in SEARCH_TERMS])
    url = (
        "http://export.arxiv.org/api/query?"
        f"search_query=({requests.utils.quote(query)})&sortBy=submittedDate&sortOrder=descending&start=0&max_results=120"
    )
    feed = feedparser.parse(url)
    papers: List[Paper] = []
    for e in feed.entries:
        published = parse_iso_datetime(e.get("published"))
        if not in_target_beijing_window(published):
            continue
        papers.append(
            Paper(
                title=e.get("title", ""),
                url=e.get("link", ""),
                abstract=sanitize_text(e.get("summary", "")),
                source="arXiv",
                published=published,
                authors=[a.name for a in e.get("authors", [])],
                institutions=[],
                author_orgs=[],
                citation_count=0,
                influence_score=0.0,
            )
        )
    return papers


def fetch_crossref() -> List[Paper]:
    from_date, to_date = beijing_day_window()
    papers: List[Paper] = []

    for term in SEARCH_TERMS:
        params = {
            "query.bibliographic": term,
            "filter": f"from-pub-date:{from_date},until-pub-date:{to_date},type:journal-article",
            "rows": 30,
            "sort": "published",
            "order": "desc",
            "select": "title,URL,abstract,container-title,author,published-online,published-print,issued,is-referenced-by-count",
        }
        resp = requests.get("https://api.crossref.org/works", params=params, timeout=30, headers=REQUEST_HEADERS)
        resp.raise_for_status()
        for item in resp.json().get("message", {}).get("items", []):
            title = (item.get("title") or [""])[0]
            abstract = html_strip(item.get("abstract") or "")
            published_online = parse_date_parts(((item.get("published-online") or {}).get("date-parts") or [[None]])[0])
            published_print = parse_date_parts(((item.get("published-print") or {}).get("date-parts") or [[None]])[0])
            issued = parse_date_parts(((item.get("issued") or {}).get("date-parts") or [[None]])[0])
            published = published_online or published_print or issued
            if not in_target_beijing_window(published):
                continue
            venue = (item.get("container-title") or ["Crossref"])[0]
            authors = [
                f"{a.get('given', '')} {a.get('family', '')}".strip()
                for a in item.get("author", [])
                if f"{a.get('given', '')} {a.get('family', '')}".strip()
            ]
            institutions = []
            author_org_pairs: List[Tuple[str, str]] = []
            for a in item.get("author", []):
                author_name = f"{a.get('given', '')} {a.get('family', '')}".strip()
                aff_names = [clean_org_name((aff.get("name") or "")) for aff in (a.get("affiliation", []) or [])]
                aff_names = [x for x in aff_names if x]
                if aff_names:
                    institutions.extend(aff_names)
                author_org_pairs.append((author_name, aff_names[0] if aff_names else ""))
            papers.append(
                Paper(
                    title=title,
                    url=item.get("URL", ""),
                    abstract=abstract,
                    source=f"Crossref/{venue}",
                    published=published,
                    authors=authors,
                    institutions=list(dict.fromkeys(institutions))[:8],
                    author_orgs=build_author_orgs(author_org_pairs),
                    citation_count=int(item.get("is-referenced-by-count") or 0),
                    influence_score=0.0,
                )
            )
    return papers


def reconstruct_abstract(inverted_index: Dict[str, List[int]]) -> str:
    if not inverted_index:
        return ""
    positions = [p for pos in inverted_index.values() for p in pos]
    if not positions:
        return ""
    words = ["" for _ in range(max(positions) + 1)]
    for w, pos in inverted_index.items():
        for p in pos:
            if 0 <= p < len(words):
                words[p] = w
    return " ".join(words).strip()


def fetch_openalex() -> List[Paper]:
    from_date, to_date = beijing_day_window()
    papers: List[Paper] = []

    for term in SEARCH_TERMS:
        params = {
            "search": term,
            "filter": f"from_publication_date:{from_date},to_publication_date:{to_date}",
            "sort": "publication_date:desc",
            "per-page": 25,
            "mailto": "report@example.com",
        }
        resp = requests.get("https://api.openalex.org/works", params=params, timeout=30, headers=REQUEST_HEADERS)
        resp.raise_for_status()
        for r in resp.json().get("results", []):
            published = parse_date_string(r.get("publication_date"))
            if not in_target_beijing_window(published):
                continue
            authors = [
                (a.get("author") or {}).get("display_name", "")
                for a in r.get("authorships", [])
                if (a.get("author") or {}).get("display_name")
            ]
            institutions = []
            author_org_pairs: List[Tuple[str, str]] = []
            for a in r.get("authorships", []):
                author_name = (a.get("author") or {}).get("display_name", "").strip()
                inst_names = [clean_org_name((inst.get("display_name") or "")) for inst in (a.get("institutions", []) or [])]
                inst_names = [x for x in inst_names if x]
                raw_aff = [clean_org_name(x) for x in (a.get("raw_affiliation_strings") or []) if isinstance(x, str)]
                raw_aff = [x for x in raw_aff if x]
                merged = inst_names or raw_aff
                if merged:
                    institutions.extend(merged)
                author_org_pairs.append((author_name, merged[0] if merged else ""))
            papers.append(
                Paper(
                    title=r.get("title") or "",
                    url=(r.get("primary_location") or {}).get("landing_page_url") or r.get("id") or "",
                    abstract=sanitize_text(reconstruct_abstract(r.get("abstract_inverted_index") or {})),
                    source="OpenAlex",
                    published=published,
                    authors=authors,
                    institutions=list(dict.fromkeys(institutions))[:8],
                    author_orgs=build_author_orgs(author_org_pairs),
                    citation_count=int(r.get("cited_by_count") or 0),
                    influence_score=float(r.get("cited_by_count") or 0) / 50.0,
                )
            )
    return papers


def fetch_semantic_scholar() -> List[Paper]:
    papers: List[Paper] = []
    base = "https://api.semanticscholar.org/graph/v1/paper/search"

    from_date, to_date = beijing_day_window()

    for term in SEARCH_TERMS:
        params = {
            "query": term,
            "limit": 25,
            "offset": 0,
            "fields": "title,abstract,url,authors.name,authors.affiliations,authors.authorId,publicationDate,publicationVenue,citationCount,influentialCitationCount",
            "year": str(now_beijing().year),
        }
        resp = requests.get(base, params=params, timeout=30, headers=REQUEST_HEADERS)
        resp.raise_for_status()
        for p in resp.json().get("data", []):
            published = parse_iso_datetime(p.get("publicationDate"))
            if not published:
                published = parse_date_string(p.get("publicationDate"))
            if not in_target_beijing_window(published):
                continue
            if not (from_date <= published.astimezone(BEIJING_TZ).strftime("%Y-%m-%d") <= to_date):
                continue
            venue = (p.get("publicationVenue") or {}).get("name") or "SemanticScholar"
            author_list = [a.get("name", "") for a in p.get("authors", []) if a.get("name")]
            author_org_pairs: List[Tuple[str, str]] = []
            institutions: List[str] = []
            for a in p.get("authors", []):
                author_name = (a.get("name") or "").strip()
                affs = [clean_org_name(aff) for aff in (a.get("affiliations") or []) if isinstance(aff, str)]
                affs = [x for x in affs if x]
                if affs:
                    institutions.extend(affs)
                author_org_pairs.append((author_name, affs[0] if affs else ""))

            papers.append(
                Paper(
                    title=p.get("title") or "",
                    url=p.get("url") or "",
                    abstract=sanitize_text(p.get("abstract") or ""),
                    source=f"SemanticScholar/{venue}",
                    published=published,
                    authors=author_list,
                    institutions=list(dict.fromkeys(institutions))[:8],
                    author_orgs=build_author_orgs(author_org_pairs),
                    citation_count=int(p.get("citationCount") or 0),
                    influence_score=float(p.get("influentialCitationCount") or 0),
                )
            )
    return papers


def fetch_rss_journals() -> List[Paper]:
    papers: List[Paper] = []
    for source_name, rss_url in RSS_SOURCES.items():
        feed = feedparser.parse(rss_url)
        for e in feed.entries:
            published = (
                parse_iso_datetime(e.get("published"))
                or parse_struct_time(e.get("published_parsed"))
            )
            if not in_target_beijing_window(published):
                continue

            title = e.get("title", "")
            summary = html_strip(e.get("summary", "") or e.get("description", ""))
            if not is_domain_relevant(title, summary):
                continue

            authors = []
            for a in e.get("authors", []):
                if getattr(a, "name", ""):
                    authors.append(a.name)
            papers.append(
                Paper(
                    title=title,
                    url=e.get("link", ""),
                    abstract=summary,
                    source=f"RSS/{source_name}",
                    published=published,
                    authors=authors,
                    institutions=[],
                    author_orgs=[],
                    citation_count=0,
                    influence_score=0.0,
                )
            )
    return papers


def dedup_rank(papers: Iterable[Paper]) -> List[Paper]:
    dedup: Dict[str, Paper] = {}
    for p in papers:
        key = normalize(p.title)
        if not key:
            continue
        prev = dedup.get(key)
        if not prev:
            dedup[key] = p
            continue
        if (topical_score(p.title, p.abstract), p.published) > (
            topical_score(prev.title, prev.abstract),
            prev.published,
        ):
            dedup[key] = p

    min_keep = max(3, int(os.environ.get("MIN_PAPER_CANDIDATES", "10")))
    strict_filtered = [p for p in dedup.values() if is_domain_relevant(p.title, p.abstract)]
    if len(strict_filtered) >= min_keep:
        strict_filtered.sort(key=lambda x: (topical_score(x.title, x.abstract), x.published), reverse=True)
        return strict_filtered

    soft_filtered = [p for p in dedup.values() if is_domain_relevant_soft(p.title, p.abstract)]
    if soft_filtered:
        print(f"[WARN] strict domain filter kept {len(strict_filtered)} papers; relaxed fallback kept {len(soft_filtered)}")
        soft_filtered.sort(key=lambda x: (topical_score(x.title, x.abstract), x.published), reverse=True)
        if len(soft_filtered) >= min_keep:
            return soft_filtered

    # final fallback: keep topical papers while still removing clearly irrelevant ones
    fallback = [
        p for p in dedup.values()
        if topical_score(p.title, p.abstract) > 0
        and not any(normalize(kw) in normalize(f"{p.title} {p.abstract}") for kw in EXCLUDED_NON_TECH_KEYWORDS)
    ]
    fallback.sort(key=lambda x: (topical_score(x.title, x.abstract), x.published), reverse=True)
    if fallback:
        print(f"[WARN] fallback topical selection enabled; candidates={len(fallback)}")
        return fallback

    strict_filtered.sort(key=lambda x: (topical_score(x.title, x.abstract), x.published), reverse=True)
    return strict_filtered




def impact_score(p: Paper) -> float:
    c = max(p.citation_count, 0)
    citation_term = math.log10(c + 1)
    return citation_term * 2.5 + max(p.influence_score, 0.0)


def diversify_sources(papers: List[Paper], limit: int) -> List[Paper]:
    """Prefer source diversity so non-arXiv papers are not starved by ranking."""
    if len(papers) <= limit:
        return papers
    by_source: Dict[str, List[Paper]] = {}
    for p in papers:
        by_source.setdefault(p.source.split("/", 1)[0], []).append(p)

    ordered_sources = sorted(by_source.keys(), key=lambda s: len(by_source[s]), reverse=True)
    for s in ordered_sources:
        by_source[s].sort(key=lambda x: (topical_score(x.title, x.abstract), x.published), reverse=True)

    out: List[Paper] = []
    while len(out) < limit:
        progressed = False
        for source in ordered_sources:
            if by_source[source]:
                out.append(by_source[source].pop(0))
                progressed = True
                if len(out) >= limit:
                    break
        if not progressed:
            break
    return out


def collect_recent_papers() -> Tuple[List[Paper], Dict[str, int]]:
    sources = [fetch_arxiv, fetch_crossref, fetch_openalex, fetch_semantic_scholar, fetch_rss_journals]
    got: List[Paper] = []
    counts: Dict[str, int] = {}

    for src in sources:
        try:
            rows = src()
            got.extend(rows)
            counts[src.__name__] = len(rows)
            print(f"[INFO] {src.__name__}: {len(rows)} rows in Beijing t-17到t-11")
        except Exception as exc:
            counts[src.__name__] = 0
            print(f"[WARN] {src.__name__} failed: {exc}")

    return dedup_rank(got), counts




def classify_paper(paper: Paper) -> str:
    txt = normalize(f"{paper.title} {paper.abstract}")
    world_keywords = [
        "world engine", "world model", "world simulator", "interactive world", "digital twin",
        "embodied", "robot simulation", "世界引擎", "世界模型", "仿真引擎",
    ]
    infra_keywords = [
        "robot data pipeline", "embodied data pipeline", "autonomy data engine", "sim2real data",
        "teleoperation data", "fleet data", "sensor data pipeline", "perception data curation",
        "multimodal robot dataset", "trajectory dataset", "具身智能数据", "机器人数据管道", "自动驾驶数据管道",
    ]
    physical_ai_keywords = ["robot", "embodied", "autonomous driving", "self-driving", "carla", "airsim", "gazebo", "具身", "机器人", "自动驾驶", "感知"]
    world_score = sum(1 for k in world_keywords if normalize(k) in txt)
    infra_score = sum(1 for k in infra_keywords if normalize(k) in txt)
    physical_score = sum(1 for k in physical_ai_keywords if normalize(k) in txt)

    if infra_score >= world_score and infra_score > 0 and physical_score > 0:
        return "Data Infra"
    if any(normalize(k) in txt for k in ["pipeline", "ingestion", "etl", "lakehouse", "feature store", "数据管道", "湖仓"]):
        return "Data Infra"
    return "World Engine"


def infer_paper_type(paper: Paper) -> str:
    src = paper.source.lower()
    if "arxiv" in src or "biorxiv" in src or "medrxiv" in src:
        return "预印本"
    if "crossref" in src or "rss" in src or "openalex" in src or "semanticscholar" in src:
        return "期刊/索引收录论文"
    return "论文"


def infer_industry_interface(paper: Paper) -> str:
    txt = normalize(f"{paper.title} {paper.abstract}")
    tags: List[str] = []
    if any(k in txt for k in ["train", "training", "policy", "训练", "学习"]):
        tags.append("训练")
    if any(k in txt for k in ["evaluate", "benchmark", "评测", "evaluation"]):
        tags.append("评测")
    if any(k in txt for k in ["simulat", "digital twin", "仿真", "world model", "world simulator"]):
        tags.append("仿真")
    if any(k in txt for k in ["deploy", "serving", "latency", "inference", "部署", "推理"]):
        tags.append("部署")
    if any(k in txt for k in ["pipeline", "etl", "ingestion", "lakehouse", "feature store", "数据管道", "湖仓"]):
        tags.append("数据管道")
    if not tags:
        tags = ["评测"]
    return " / ".join(dict.fromkeys(tags))


def relevance_components(paper: Paper) -> Tuple[float, float, float]:
    txt = normalize(f"{paper.title} {paper.abstract}")
    relevance = float(max(topical_score(paper.title, paper.abstract), 1))

    method_cues = ["method", "approach", "framework", "architecture", "算法", "方法", "框架"]
    result_cues = ["result", "improv", "benchmark", "实验", "结果", "提升", "accuracy"]
    info_increment = 1.0
    if paper.abstract:
        info_increment += min(len(paper.abstract) / 1200.0, 1.5)
    info_increment += 0.4 * sum(1 for k in method_cues if normalize(k) in txt)
    info_increment += 0.4 * sum(1 for k in result_cues if normalize(k) in txt)

    industry_value = 1.0
    interface = infer_industry_interface(paper)
    industry_value += 0.5 * interface.count("/")
    industry_value += min(impact_score(paper) / 8.0, 2.0)
    return relevance, info_increment, industry_value


def ranking_score(paper: Paper) -> float:
    r, i, v = relevance_components(paper)
    return r * i * v


def _parse_arxiv_id(url: str) -> str:
    m = re.search(r"arxiv\.org/(abs|pdf)/([0-9]{4}\.[0-9]{4,5})(v\d+)?", url or "", re.I)
    return m.group(2) if m else ""


def _detect_links(text: str) -> Dict[str, str]:
    t = text or ""
    out = {"github": "", "huggingface": "", "paperswithcode": ""}
    g = re.search(r"https?://github\.com/[\w\-\.]+/[\w\-\.]+", t, re.I)
    h = re.search(r"https?://huggingface\.co/[\w\-\./]+", t, re.I)
    pwc = re.search(r"https?://paperswithcode\.com/[\w\-\./]+", t, re.I)
    if g:
        out["github"] = g.group(0)
    if h:
        out["huggingface"] = h.group(0)
    if pwc:
        out["paperswithcode"] = pwc.group(0)
    return out


def _github_repo_slug(repo_url: str) -> Tuple[str, str]:
    m = re.search(r"github\.com/([\w\-\.]+)/([\w\-\.]+)", repo_url or "")
    if not m:
        return "", ""
    return m.group(1), m.group(2).replace('.git', '')


def _github_auth_headers(extra: Optional[Dict[str, str]] = None) -> Dict[str, str]:
    token = os.environ.get("GITHUB_TOKEN", "").strip()
    headers = dict(REQUEST_HEADERS)
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if extra:
        headers.update(extra)
    return headers


def _github_repo_snapshot(repo_url: str) -> Dict[str, object]:
    owner, repo = _github_repo_slug(repo_url)
    if not owner or not repo:
        return {}
    try:
        r = requests.get(
            f"https://api.github.com/repos/{owner}/{repo}",
            timeout=15,
            headers=_github_auth_headers(),
        )
        if r.status_code != 200:
            return {}
        return r.json() or {}
    except Exception:
        return {}


def _github_repo_metrics(repo_url: str) -> Dict[str, float]:
    owner, repo = _github_repo_slug(repo_url)
    if not owner or not repo:
        return {
            "open_source_base": 0.0,
            "star_velocity": 0.0,
            "deep_engagement": 0.0,
            "github_total": 0.0,
            "stars": 0.0,
            "forks": 0.0,
            "issues": 0.0,
            "prs": 0.0,
            "star_events_7d": 0.0,
        }

    snapshot = _github_repo_snapshot(repo_url)
    stars = float(snapshot.get("stargazers_count") or 0)
    forks = float(snapshot.get("forks_count") or 0)
    open_issues = float(snapshot.get("open_issues_count") or 0)
    description = str(snapshot.get("description") or "").strip()

    # 1) Open-source baseline (10)
    open_source_base = 2.0
    readme_ok, core_code_ok = False, False
    try:
        c = requests.get(
            f"https://api.github.com/repos/{owner}/{repo}/contents",
            timeout=15,
            headers=_github_auth_headers(),
        )
        if c.status_code == 200:
            items = c.json() or []
            names = {str(x.get("name", "")).lower() for x in items if isinstance(x, dict)}
            readme_ok = any(n.startswith("readme") for n in names)
            core_code_ok = any(n in names for n in {"src", "train.py", "inference.py", "model", "models", "scripts", "requirements.txt"})
    except Exception:
        pass
    open_source_base += 4.0 if readme_ok else 0.0
    open_source_base += 4.0 if core_code_ok else 0.0
    if description:
        open_source_base += 0.5
    open_source_base = min(10.0, open_source_base)

    # 2) Star velocity (15): count recent WatchEvent over last 7 days.
    star_events_7d = 0.0
    since = dt.datetime.now(dt.timezone.utc) - dt.timedelta(days=7)
    try:
        e = requests.get(
            f"https://api.github.com/repos/{owner}/{repo}/events",
            params={"per_page": 100},
            timeout=15,
            headers=_github_auth_headers(),
        )
        if e.status_code == 200:
            events = e.json() or []
            for it in events:
                if str(it.get("type", "")) != "WatchEvent":
                    continue
                ts = str(it.get("created_at", ""))
                try:
                    d = dt.datetime.fromisoformat(ts.replace("Z", "+00:00"))
                    if d >= since:
                        star_events_7d += 1.0
                except Exception:
                    continue
    except Exception:
        pass
    star_velocity = min(15.0, 5.5 * math.log1p(star_events_7d))

    # 3) Deep engagement (15): forks + effective issues + PRs
    prs = 0.0
    issues_eff = max(0.0, open_issues * 0.65)
    try:
        q_since = since.strftime("%Y-%m-%d")
        q = f"repo:{owner}/{repo} type:pr created:>={q_since}"
        sr = requests.get(
            "https://api.github.com/search/issues",
            params={"q": q, "per_page": 1},
            timeout=15,
            headers=_github_auth_headers(),
        )
        if sr.status_code == 200:
            prs = float(sr.json().get("total_count") or 0)
    except Exception:
        pass
    deep_signal = forks + issues_eff + prs
    deep_engagement = min(15.0, 3.8 * math.log1p(deep_signal))

    total = min(40.0, open_source_base + star_velocity + deep_engagement)
    return {
        "open_source_base": round(open_source_base, 2),
        "star_velocity": round(star_velocity, 2),
        "deep_engagement": round(deep_engagement, 2),
        "github_total": round(total, 2),
        "stars": round(stars, 2),
        "forks": round(forks, 2),
        "issues": round(open_issues, 2),
        "prs": round(prs, 2),
        "star_events_7d": round(star_events_7d, 2),
    }


def _github_metrics(repo_url: str) -> Tuple[int, int]:
    # Backward-compatible shortcut for existing early quality scorer.
    m = _github_repo_metrics(repo_url)
    return int(m.get("stars") or 0), int(m.get("forks") or 0)


def _x_discussion_score(query: str) -> Dict[str, float]:
    if not query.strip():
        return {"kol_score": 0.0, "interaction_score": 0.0, "x_total": 0.0, "mentions": 0.0}

    # optional richer path with X API bearer token
    bearer = os.environ.get("X_BEARER_TOKEN", "").strip()
    if bearer:
        try:
            r = requests.get(
                "https://api.x.com/2/tweets/search/recent",
                params={
                    "query": f'"{query}" lang:en -is:retweet',
                    "max_results": 30,
                    "tweet.fields": "public_metrics,author_id",
                    "expansions": "author_id",
                    "user.fields": "public_metrics,verified",
                },
                headers={"Authorization": f"Bearer {bearer}"},
                timeout=15,
            )
            if r.status_code == 200:
                js = r.json() or {}
                tweets = js.get("data") or []
                users = {u.get("id"): u for u in ((js.get("includes") or {}).get("users") or [])}
                likes = quotes = 0.0
                kol_mass = 0.0
                for t in tweets:
                    pm = t.get("public_metrics") or {}
                    likes += float(pm.get("like_count") or 0)
                    quotes += float(pm.get("quote_count") or 0)
                    uid = t.get("author_id")
                    u = users.get(uid) or {}
                    followers = float(((u.get("public_metrics") or {}).get("followers_count") or 0))
                    verified_bonus = 1.35 if u.get("verified") else 1.0
                    kol_mass += verified_bonus * math.log1p(max(0.0, followers))
                kol_score = min(15.0, 1.8 * math.log1p(kol_mass))
                interaction_score = min(15.0, 1.9 * math.log1p(3 * quotes + likes))
                total = min(30.0, kol_score + interaction_score)
                return {
                    "kol_score": round(kol_score, 2),
                    "interaction_score": round(interaction_score, 2),
                    "x_total": round(total, 2),
                    "mentions": float(len(tweets)),
                }
        except Exception:
            pass

    # fallback: RSS mention count + weak KOL heuristic from handle list
    endpoints = os.environ.get(
        "X_SEARCH_RSS_ENDPOINTS",
        "https://nitter.net/search/rss?q={q},https://nitter.poast.org/search/rss?q={q}",
    ).split(",")
    entries = []
    for ep in endpoints:
        ep = ep.strip()
        if not ep:
            continue
        try:
            url = ep.format(q=requests.utils.quote(query))
            feed = feedparser.parse(url)
            entries = getattr(feed, "entries", []) or []
            if entries:
                break
        except Exception:
            continue
    mentions = float(len(entries))
    kol_handles = {
        "ylecun": 2.0, "karpathy": 2.0, "goodfellow_ian": 2.0, "demishassabis": 1.8,
        "andrewng": 1.8, "sama": 1.6, "gdb": 1.5,
    }
    kol_hit = 0.0
    for e in entries:
        txt = f"{str(getattr(e, 'title', '') or '')} {str(getattr(e, 'author', '') or '')}".lower()
        for h, w in kol_handles.items():
            if h in txt:
                kol_hit += w
    kol_score = min(15.0, 5.0 * math.log1p(kol_hit))
    interaction_score = min(15.0, 3.0 * math.log1p(max(0.0, mentions)))
    total = min(30.0, kol_score + interaction_score)
    return {
        "kol_score": round(kol_score, 2),
        "interaction_score": round(interaction_score, 2),
        "x_total": round(total, 2),
        "mentions": round(mentions, 2),
    }


def _reddit_discussion_score(query: str) -> Dict[str, float]:
    if not query.strip():
        return {
            "vertical_score": 0.0,
            "depth_consensus": 0.0,
            "reddit_total": 0.0,
            "posts": 0.0,
        }
    try:
        r = requests.get(
            "https://www.reddit.com/search.json",
            params={"q": query, "sort": "new", "limit": 30, "restrict_sr": "false"},
            timeout=15,
            headers={**REQUEST_HEADERS, "User-Agent": "Mozilla/5.0 (daily-paper-agent)"},
        )
        if r.status_code != 200:
            return {"vertical_score": 0.0, "depth_consensus": 0.0, "reddit_total": 0.0, "posts": 0.0}
        children = ((r.json().get("data") or {}).get("children") or [])
        posts = [c.get("data") or {} for c in children]
        if not posts:
            return {"vertical_score": 0.0, "depth_consensus": 0.0, "reddit_total": 0.0, "posts": 0.0}

        sub_weights = {
            "machinelearning": 2.0,
            "localllama": 1.8,
            "singularity": 1.6,
            "robotics": 1.8,
            "mlscaling": 1.7,
            "artificial": 1.4,
            "technology": 1.0,
        }
        weighted_posts = 0.0
        comments_sum = 0.0
        upvote_quality = 0.0
        for p in posts:
            sub = str(p.get("subreddit") or "").lower()
            w = sub_weights.get(sub, 1.0)
            weighted_posts += w
            comments_sum += w * float(p.get("num_comments") or 0)
            upvote_quality += w * (float(p.get("score") or 0) * float(p.get("upvote_ratio") or 0))

        vertical_score = min(10.0, 2.8 * math.log1p(weighted_posts))
        depth_consensus = min(20.0, 2.2 * math.log1p(comments_sum) + 1.4 * math.log1p(max(0.0, upvote_quality)))
        total = min(30.0, vertical_score + depth_consensus)
        return {
            "vertical_score": round(vertical_score, 2),
            "depth_consensus": round(depth_consensus, 2),
            "reddit_total": round(total, 2),
            "posts": float(len(posts)),
        }
    except Exception:
        return {"vertical_score": 0.0, "depth_consensus": 0.0, "reddit_total": 0.0, "posts": 0.0}


def compute_social_discussion_score(paper: Paper) -> Tuple[float, Dict[str, object]]:
    """100-point importance score: GitHub 40 + X 30 + Reddit 30."""
    arxiv_id = _parse_arxiv_id(paper.url)
    title_query = re.sub(r"\s+", " ", paper.title).strip()[:180]
    query = arxiv_id or title_query

    text_blob = "\n".join([paper.title, paper.abstract])
    links = _detect_links(text_blob)
    gh = _github_repo_metrics(links.get("github", ""))
    x = _x_discussion_score(query)
    rd = _reddit_discussion_score(query)

    total = float(gh.get("github_total", 0.0)) + float(x.get("x_total", 0.0)) + float(rd.get("reddit_total", 0.0))
    total = round(min(100.0, total), 2)
    details: Dict[str, object] = {
        "github": gh,
        "x": x,
        "reddit": rd,
        "importance_score": total,
        "query": query,
    }
    return total, details


def pick_top_discussed_papers(papers: List[Paper], limit: int = 3) -> List[Paper]:
    if not papers:
        return []
    pool_size = int(os.environ.get("DISCUSSION_CANDIDATE_POOL", "24"))
    pool = sorted(papers, key=ranking_score, reverse=True)[: max(limit, pool_size)]

    scored: List[Tuple[Paper, float]] = []
    for p in pool:
        social_score, details = compute_social_discussion_score(p)
        setattr(p, "_social_score", social_score)
        setattr(p, "_social_details", details)
        scored.append((p, social_score))

    scored.sort(key=lambda x: (x[1], ranking_score(x[0]), topical_score(x[0].title, x[0].abstract)), reverse=True)
    return [x[0] for x in scored[:limit]]


def compute_early_quality_score(paper: Paper, category: str, fulltext_context: str) -> Dict[str, object]:
    text_blob = "\n".join([paper.title, paper.abstract, fulltext_context])
    links = _detect_links(text_blob)
    github_stars, github_forks = _github_metrics(links["github"])

    # A. Author Historical Hit Rate (conservative if data missing)
    core_authors = paper.authors[:3]
    relevant_last4y = 0
    high_impact_last4y = 0
    hit_rate = 0.0
    a_score = 0
    a_reason = "作者历史命中数据缺失，按保守分处理。"

    # B. Reproducibility & Artifact Completeness
    low = normalize(text_blob)
    has_repo = bool(links["github"])
    has_instr = any(k in low for k in ["installation", "usage", "train", "inference", "quick start", "运行", "训练", "推理"])
    has_weights = any(k in low for k in ["pretrained", "checkpoint", "weights", "模型权重"])
    has_dataset_or_eval = any(k in low for k in ["dataset", "benchmark", "evaluation script", "数据集", "评测脚本"])
    has_license = any(k in low for k in ["license", "mit", "apache-2.0", "bsd"])
    has_env = any(k in low for k in ["requirements.txt", "environment.yml", "dockerfile", "conda", "poetry"])
    b_score = min(25, (8 if has_repo else 0) + (4 if has_instr else 0) + (5 if has_weights else 0) + (4 if has_dataset_or_eval else 0) + (2 if has_license else 0) + (2 if has_env else 0))

    # C. Early Community Response
    c_score = 0
    if github_stars >= 1000:
        c_score += 12
    elif github_stars >= 300:
        c_score += 9
    elif github_stars >= 100:
        c_score += 6
    elif github_stars >= 30:
        c_score += 3
    elif github_stars >= 1:
        c_score += 1
    if github_forks >= 100:
        c_score += 4
    elif github_forks >= 30:
        c_score += 3
    elif github_forks >= 10:
        c_score += 2
    elif github_forks >= 1:
        c_score += 1
    pwc_listed = bool(links["paperswithcode"])
    hf_signal = bool(links["huggingface"])
    c_score += 2 if pwc_listed else 0
    c_score += 2 if hf_signal else 0
    c_score = min(20, c_score)

    # D. Experimental Strength
    strong_baselines = any(k in low for k in ["baseline", "sota", "state-of-the-art", "对比方法", "基线"])
    ablation = any(k in low for k in ["ablation", "消融"])
    multi_task = any(k in low for k in ["multi-task", "multiple datasets", "environments", "多个数据集", "多任务", "多环境"])
    efficiency = any(k in low for k in ["latency", "flops", "memory", "throughput", "training cost", "inference cost", "时延", "吞吐", "显存", "成本"])
    limitations = any(k in low for k in ["limitation", "future work", "局限", "未来工作"])
    benchmark_tables = any(k in low for k in ["table", "benchmark", "leaderboard", "表", "基准"])
    d_score = min(15, (4 if strong_baselines else 0) + (3 if ablation else 0) + (3 if multi_task else 0) + (2 if efficiency else 0) + (1 if limitations else 0) + (2 if benchmark_tables else 0))

    # E. Novelty & Problem Importance
    problem_clarity = 3 if any(k in low for k in ["we address", "problem", "challenge", "我们解决", "问题", "挑战"]) else 1
    contribution_specificity = 3 if any(k in low for k in ["we propose", "introduce", "our method", "提出", "方法", "框架"]) else 1
    difference_prior = 2 if any(k in low for k in ["compared with", "different from", "prior work", "相比", "不同于", "已有方法"]) else 1
    frontier = 2 if any(k in low for k in ["world model", "coding agent", "foundation model", "multimodal", "robot", "physical ai", "data infra", "世界模型", "具身", "机器人"]) else 1
    e_score = min(10, problem_clarity + contribution_specificity + difference_prior + frontier)

    total = int(max(0, min(100, a_score + b_score + c_score + d_score + e_score)))

    # confidence
    confidence = 1.0
    if relevant_last4y == 0:
        confidence -= 0.15
    if not has_repo:
        confidence -= 0.15
    if not (links["github"] or links["paperswithcode"] or links["huggingface"]):
        confidence -= 0.10
    if not has_readable_fulltext(fulltext_context):
        confidence -= 0.10
    confidence = max(0.3, round(confidence, 2))

    days_since = max(0, (now_beijing().date() - paper.published.astimezone(BEIJING_TZ).date()).days)
    arxiv_id = _parse_arxiv_id(paper.url)
    availability = {
        "arxiv": "arxiv" in (paper.source or "").lower() or "arxiv.org" in (paper.url or "").lower(),
        "semantic_scholar": "semanticscholar" in (paper.source or "").lower(),
        "openalex": "openalex" in (paper.source or "").lower(),
        "papers_with_code": bool(links["paperswithcode"]),
        "github": bool(links["github"]),
        "huggingface": bool(links["huggingface"]),
    }

    missing = [k for k, v in availability.items() if not v]
    tier = "A" if total >= 85 else "B" if total >= 70 else "C" if total >= 50 else "D"
    label = {"A": "very strong early signal", "B": "promising", "C": "mixed evidence", "D": "weak/insufficient"}[tier]

    return {
      "paper": {
        "title": paper.title,
        "arxiv_id": arxiv_id,
        "arxiv_url": f"https://arxiv.org/abs/{arxiv_id}" if arxiv_id else "",
        "published_date": paper.published.astimezone(BEIJING_TZ).strftime("%Y-%m-%d"),
        "days_since_release": days_since,
        "primary_category": category,
        "authors": core_authors,
      },
      "data_availability": availability,
      "scores": {
        "author_historical_hit_rate": {
          "score": a_score, "max_score": 30,
          "raw_metrics": {
            "core_authors": core_authors,
            "relevant_papers_last_4y": relevant_last4y,
            "high_impact_papers_last_4y": high_impact_last4y,
            "hit_rate": hit_rate,
          },
          "reasoning": a_reason,
        },
        "reproducibility_artifacts": {
          "score": b_score, "max_score": 25,
          "raw_metrics": {
            "code_repo": has_repo,
            "instructions": has_instr,
            "weights": has_weights,
            "dataset_or_eval_scripts": has_dataset_or_eval,
            "license": has_license,
            "env_files": has_env,
          },
          "reasoning": "依据代码/权重/说明文档等可复现资产打分。",
        },
        "early_community_response": {
          "score": c_score, "max_score": 20,
          "raw_metrics": {
            "github_stars": github_stars,
            "github_forks": github_forks,
            "papers_with_code_listed": pwc_listed,
            "huggingface_signal": hf_signal,
          },
          "reasoning": "依据早期社区可观测信号打分；新发布论文不做过度解读。",
        },
        "experimental_strength": {
          "score": d_score, "max_score": 15,
          "raw_metrics": {
            "strong_baselines": strong_baselines,
            "ablation": ablation,
            "multi_dataset_or_multi_task": multi_task,
            "efficiency_metrics": efficiency,
            "limitations": limitations,
            "benchmark_tables": benchmark_tables,
          },
          "reasoning": "依据正文中的实验设计与报告完整度打分。",
        },
        "novelty_and_problem_importance": {
          "score": e_score, "max_score": 10,
          "raw_metrics": {
            "problem_clarity": problem_clarity,
            "contribution_specificity": contribution_specificity,
            "difference_from_prior_work": difference_prior,
            "frontier_relevance": frontier,
          },
          "reasoning": "依据问题清晰度、贡献具体性与前沿相关性打分。",
        },
        "total_score": total,
        "confidence": confidence,
      },
      "verdict": {
        "tier": tier,
        "label": label,
        "summary": "早期质量评分用于快速优先级排序，需结合后续复现实验继续验证。",
        "main_strengths": ["结构化实验信号", "可复现资产信号"],
        "main_risks": ["作者历史数据缺失", "早期社区信号波动"],
      },
      "implementation_notes": {
        "missing_data": missing,
        "assumptions": ["仅使用可程序化访问数据源", "缺失字段按保守分处理"],
        "recommended_followup_checks": ["补全作者历史论文画像", "人工复核关键实验表格"],
      }
    }


def build_day_summary(papers: List[Paper]) -> str:
    total = len(papers)
    world = sum(1 for p in papers if classify_paper(p) == "World Engine")
    infra = total - world
    return f"本周总篇数：{total}；World Engine：{world}；Data Infra：{infra}"




PAPERS_DIR = pathlib.Path(os.environ.get("PAPERS_DIR", "papers"))


def _extract_arxiv_id(url: str) -> Optional[str]:
    """Extract arXiv paper ID from various URL formats."""
    if not url:
        return None
    m = re.search(r"arxiv\.org/(?:abs|pdf|html)/(\d{4}\.\d{4,5}(?:v\d+)?)", url)
    return m.group(1) if m else None


def download_pdf(paper: "Paper", dest_dir: pathlib.Path | None = None, max_retries: int = 3) -> Optional[pathlib.Path]:
    """Download the PDF for a paper. Returns the local file path, or None on failure.

    Strategy:
      1. arXiv papers  -> https://arxiv.org/pdf/{id}.pdf
      2. Other papers   -> try the paper URL directly if it ends with .pdf

    Retries each URL up to *max_retries* times with exponential backoff to
    handle transient failures and arXiv rate-limiting (HTTP 429 / 503).
    """
    dest_dir = dest_dir or PAPERS_DIR
    dest_dir.mkdir(parents=True, exist_ok=True)

    arxiv_id = _extract_arxiv_id(paper.url or "")

    pdf_urls: List[str] = []
    if arxiv_id:
        pdf_urls.append(f"https://arxiv.org/pdf/{arxiv_id}.pdf")
    if (paper.url or "").lower().endswith(".pdf"):
        pdf_urls.append(paper.url)

    if not pdf_urls:
        print(f"[PDF] no downloadable URL for: {(paper.title or '')[:60]}")
        return None

    safe_name = re.sub(r"[^\w\-.]", "_", (paper.title or "paper")[:80]).strip("_")
    dest_path = dest_dir / f"{safe_name}.pdf"

    for url in pdf_urls:
        for attempt in range(1, max_retries + 1):
            try:
                resp = requests.get(url, timeout=60, headers=REQUEST_HEADERS, stream=True)
                if resp.status_code in (429, 503):
                    wait = 2 ** attempt
                    print(f"[PDF] rate-limited ({resp.status_code}), retry {attempt}/{max_retries} after {wait}s: {url}")
                    import time; time.sleep(wait)
                    continue
                resp.raise_for_status()
                content_type = resp.headers.get("Content-Type", "")
                if "pdf" not in content_type and "octet-stream" not in content_type:
                    print(f"[PDF] unexpected content-type '{content_type}' from {url}, skipping")
                    break  # try next URL, not next retry
                with open(dest_path, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=64 * 1024):
                        f.write(chunk)
                if dest_path.stat().st_size > 1024:
                    print(f"[PDF] downloaded: {dest_path.name} ({dest_path.stat().st_size // 1024}KB)")
                    return dest_path
                else:
                    dest_path.unlink(missing_ok=True)
                    print(f"[PDF] file too small (<1KB), discarded: {dest_path.name}")
                    break
            except Exception as exc:
                wait = 2 ** attempt
                print(f"[PDF] attempt {attempt}/{max_retries} failed for {url}: {exc}")
                if attempt < max_retries:
                    import time; time.sleep(wait)
                continue
    print(f"[PDF] all attempts exhausted for: {(paper.title or '')[:60]}")
    return None


def extract_text_from_pdf(pdf_path: pathlib.Path, max_chars: int = 60000) -> str:
    """Extract text from a PDF file using PyMuPDF (fitz)."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        print("[PDF] PyMuPDF not installed, skipping PDF text extraction")
        return ""

    try:
        doc = fitz.open(str(pdf_path))
        texts: List[str] = []
        total = 0
        for page in doc:
            page_text = page.get_text("text")
            texts.append(page_text)
            total += len(page_text)
            if total >= max_chars:
                break
        doc.close()
        full = "\n".join(texts)
        return full[:max_chars]
    except Exception as exc:
        print(f"[PDF] text extraction failed for {pdf_path}: {exc}")
        return ""


def fetch_fulltext_via_pdf(paper: "Paper", dest_dir: pathlib.Path | None = None) -> Tuple[str, Optional[pathlib.Path]]:
    """Download PDF, extract text. Returns (extracted_text, pdf_path)."""
    pdf_path = download_pdf(paper, dest_dir=dest_dir)
    if pdf_path is None:
        return "", None
    text = extract_text_from_pdf(pdf_path)
    return text, pdf_path


def fetch_fulltext_context(paper: Paper) -> str:
    """Fallback: fetch full text via HTML scraping (used when PDF is unavailable)."""
    candidates = [paper.url]
    if "arxiv.org/abs/" in (paper.url or ""):
        candidates.append((paper.url or "").replace("/abs/", "/html/"))

    best = ""
    for u in candidates:
        if not u:
            continue
        try:
            resp = requests.get(u, timeout=20, headers=REQUEST_HEADERS)
            resp.raise_for_status()
            txt = sanitize_text(resp.text, max_len=20000)
            if len(txt) > len(best):
                best = txt
        except Exception:
            continue
    return best


def has_readable_fulltext(fulltext_context: str) -> bool:
    # protect against abstract-only snippets / nav noise; configurable lower bound for real-world sites
    min_chars = int(os.environ.get("FULLTEXT_MIN_CHARS", "1200"))
    return len((fulltext_context or "").strip()) >= min_chars


def _build_social_buzz_context(paper: Paper) -> str:
    """Build a human-readable summary of why this paper is getting attention."""
    details = getattr(paper, "_social_details", None)
    if not details:
        return ""
    parts = []
    gh = details.get("github") or {}
    gh_total = float(gh.get("github_total", 0))
    if gh_total > 5:
        stars = gh.get("stargazers_count") or gh.get("stars_count") or gh.get("stars", 0)
        forks = gh.get("forks_count") or gh.get("forks", 0)
        if stars:
            parts.append(f"GitHub 项目已获 {stars} stars" + (f", {forks} forks" if forks else ""))
        elif gh_total > 10:
            parts.append("在 GitHub 上受到开发者社区关注")

    x = details.get("x") or {}
    x_total = float(x.get("x_total", 0))
    kol_names = [str(n) for n in (x.get("kol_names") or x.get("kol_list") or []) if n]
    if kol_names:
        parts.append(f"被 {', '.join(kol_names[:3])} 等业内大佬在 X 上转发讨论")
    elif x_total > 5:
        parts.append("在 X (Twitter) 上引发技术社区讨论")

    rd = details.get("reddit") or {}
    rd_total = float(rd.get("reddit_total", 0))
    subreddits = [str(s) for s in (rd.get("subreddits") or rd.get("top_subreddits") or []) if s]
    if subreddits:
        parts.append(f"在 Reddit r/{subreddits[0]} 等社区被热议")
    elif rd_total > 5:
        parts.append("在 Reddit 机器学习社区获得关注")

    return "；".join(parts) if parts else ""


def build_prompt(paper: Paper, category: str, fulltext_context: str) -> str:
    social_buzz = _build_social_buzz_context(paper)
    social_hint = ""
    if social_buzz:
        social_hint = f"\n        社区反响参考（仅供撰写「为什么值得关注」时参考）：{social_buzz}"

    return textwrap.dedent(
        f"""
        你是一位资深科技行业分析师，正在为科技公司高管撰写前沿技术研究简报。
        读者是普通大学毕业的科技企业高管，不是AI/数据/机器人领域的专家。

        ── 核心原则：让外行高管读懂 ──
        1) 目标读者画像：科技公司VP/CTO级别，聪明但不深入具体技术细节。
           他们关心的是：这个技术解决什么业务问题？比现有方案好在哪？对我们公司有什么影响？
        2) 写法参照券商/咨询公司的行业研究报告，而非学术论文摘要。
           先讲清楚「是什么」和「解决什么问题」，再讲「怎么做到的」。
        3) 【去术语化】遇到专业术语必须用日常语言解释清楚，不能假设读者知道。
           × 「采用自回归范式替代传统匹配式跟踪」
           → ✓ 「用逐步预测的方式（类似语言模型逐字生成文本）来追踪物体运动，取代了传统逐帧匹配的方法」
           × 「在nuScenes基准上AMOTA绝对提升20.2个百分点」
           → ✓ 「在自动驾驶领域最主流的测试集上，追踪准确率提升了约20个百分点——这是一个非常显著的进步」
           × 「通过扁平化高斯基元优化表面贴合度」
           → ✓ 「通过把3D模型的基本构建单元从球形改为薄片形，让模型表面更贴合真实物体」
        4) 【禁止学术腔】不得出现「论文提出/报告/声称/指出」「作者认为/发现」「本文/该研究」等学术转述句式。
           直接以技术或产品为主语陈述事实。
        5) 每段先给结论，再用一句话补关键证据。如果原文缺失某项信息，写「原文未披露」。
        6) 每一行必须是完整句，不能半句收尾。

        严格输出以下六行，不要输出其它字段：
        为什么值得关注：<用1-2句话，用轻松但有说服力的语气告诉读者这篇为什么重要——可以提到它在GitHub上多受欢迎、被哪些业内知名人物转发讨论、在Reddit社区引发了什么样的讨论，让读者感受到「这篇是真的有人在意」。如果没有明显的社区热度，就从技术突破的角度说明其重要性>
        问题与背景：<用2-3句话讲清楚这个技术要解决什么现实问题，现有方案有什么不足，让非专业人士也能理解>
        核心方法与创新：<用2-3句话用通俗语言解释技术思路，遇到专业概念必须括号注释，重点讲「做了什么」而非罗列技术名词>
        关键结论：<用2-3句话讲效果，数据要给出直观对比和解读，不要只列数字不解释含义>
        增量价值与影响：<用2-3句话讲这项技术对行业或企业的实际意义，它改变了什么，谁会受益>
        局限与开放问题：<用2-3句话讲当前的不足和未解决的问题，以及距离实际落地还差什么>

        论文标题：{paper.title}
        分类：{category}
        作者：{", ".join(paper.authors[:10]) if paper.authors else "未披露"}
        摘要：{paper.abstract[:3000] if paper.abstract else "未披露"}{social_hint}

        以下是正文提取片段（若为空表示仅抓到摘要）：
        {fulltext_context[:30000] if fulltext_context else "未抓取到正文"}
        """
    ).strip()


def analyze_paper(client: OpenAI, paper: Paper, category: str, fulltext_context: str) -> str:
    completion = client.chat.completions.create(
        model=os.environ.get("GEMINI_MODEL", DEFAULT_MODEL),
        temperature=0.1,
        max_tokens=65536,
        messages=[
            {"role": "system", "content": "你是资深科技行业分析师，为科技公司高管撰写技术研究简报。读者是聪明但非技术专家的企业高管，所有专业术语必须用通俗语言解释。行文简洁有力，直接陈述事实与判断，绝不使用'论文指出/报告/声称'等学术转述句式。"},
            {"role": "user", "content": build_prompt(paper, category, fulltext_context)},
        ],
    )
    return (completion.choices[0].message.content or "").strip()




def clean_symbols(text: str) -> str:
    cleaned = text.replace("#", "").replace("*", "")
    lines = [re.sub(r"^[\-\s]+", "", line) for line in cleaned.splitlines()]
    return "\n".join(lines)


def parse_structured_analysis(text: str) -> Dict[str, str]:
    keys = [
        "为什么值得关注",
        "问题与背景",
        "核心方法与创新",
        "关键结论",
        "增量价值与影响",
        "局限与开放问题",
    ]
    aliases = {
        "为什么值得关注": "为什么值得关注",
        "为何值得关注": "为什么值得关注",
        "关注理由": "为什么值得关注",
        "问题与背景": "问题与背景",
        "论文想解决什么问题、该问题为什么重要": "问题与背景",
        "问题与重要性": "问题与背景",
        "核心方法与创新": "核心方法与创新",
        "论文的核心方法是什么、和以前相比如何创新": "核心方法与创新",
        "关键结论": "关键结论",
        "论文的核心结论": "关键结论",
        "核心结论": "关键结论",
        "增量价值与影响": "增量价值与影响",
        "论文的增量价值是什么、会带来什么影响": "增量价值与影响",
        "局限与开放问题": "局限与开放问题",
        "论文的局限性和不确定性、没有解决什么问题": "局限与开放问题",
        "局限性和不确定性": "局限与开放问题",
    }

    data: Dict[str, str] = {k: "未披露" for k in keys}
    for line in clean_symbols(text).splitlines():
        sline = line.strip()
        if "：" not in sline:
            continue
        k, v = sline.split("：", 1)
        k = aliases.get(k.strip(), "")
        if k and v.strip():
            data[k] = v.strip()

    for k in keys:
        if data[k].strip() in ["未披露", "摘要未披露", "未知", "不详"]:
            data[k] = "原文未明确说明。"
        data[k] = _keep_first_sentences(_finalize_sentence(data[k]), 3)
    return data


def _finalize_sentence(text: str) -> str:
    t = re.sub(r"\s+", " ", (text or "")).replace("…", "").strip()
    if not t:
        return "未披露。"
    t = re.sub(r"(和|及|与|并|以及|并且|等)\s*[。！？]$", "。", t)
    t = re.sub(r"，[^，。！？；]{1,2}[。！？]$", "。", t)
    if not re.search(r"[。！？]$", t):
        t += "。"
    return t


def _keep_first_sentences(text: str, max_sentences: int) -> str:
    t = _finalize_sentence(text)
    pieces = [x.strip() for x in re.split(r"(?<=[。！？])", t) if x.strip()]
    if not pieces:
        return t
    out = "".join(pieces[:max_sentences]).strip()
    return _finalize_sentence(out)


def _trim_complete(text: str, max_len: int) -> str:
    t = _finalize_sentence(text)
    if len(t) <= max_len:
        return t
    window = t[:max_len]
    m = max(window.rfind("。"), window.rfind("！"), window.rfind("？"), window.rfind("；"))
    if m >= int(max_len * 0.5):
        return window[:m + 1]
    return _finalize_sentence(re.sub(r"(和|及|与|并|以及|并且|等)$", "", window[: max_len - 1].rstrip("，、；： ")) )


def confidence_level(paper: Paper) -> str:
    if paper.abstract and len(paper.abstract) > 900:
        return "高"
    if paper.abstract and len(paper.abstract) > 300:
        return "中"
    return "低"





def format_author_orgs(paper: Paper) -> str:
    if paper.author_orgs:
        return "，".join(paper.author_orgs[:6])
    authors = paper.authors[:6] if paper.authors else []
    insts = paper.institutions[:6] if paper.institutions else []
    if authors and insts:
        return f"{', '.join(authors)}（{';'.join(insts)}）"
    if authors:
        return ", ".join(authors)
    if insts:
        return "（" + ";".join(insts) + "）"
    return "未披露"


def render_paper_block(index: int, item: AnalyzedPaper, parsed: Dict[str, str], rank_pos: int) -> List[str]:
    paper = item.paper
    domain_tag = "世界模型 / 仿真系统" if item.category == "World Engine" else "数据基础设施 / 数据工程"
    published_bj = paper.published.astimezone(BEIJING_TZ).strftime("%Y-%m-%d %H:%M")
    return [
        "分隔线",
        f"标题：[{domain_tag}] {paper.title}",
        f"发布时间：{published_bj}（北京时间）",
        f"链接：{paper.url}",
        f"作者：{format_author_orgs(paper)}",
        f"为什么值得关注：{parsed['为什么值得关注']}",
        f"问题与背景：{parsed['问题与背景']}",
        f"核心方法与创新：{parsed['核心方法与创新']}",
        f"关键结论：{parsed['关键结论']}",
        f"增量价值与影响：{parsed['增量价值与影响']}",
        f"局限与开放问题：{parsed['局限与开放问题']}",
    ]


def fallback_structured_analysis(paper: Paper, category: str, reason: str = "") -> Dict[str, str]:
    abstract = sanitize_text((paper.abstract or ""), max_len=1200)
    short_abstract = abstract if abstract else ""
    no_data_msg = "本条未能抓取到稳定正文，以下结论基于题目与摘要，需后续复核。"
    reason_text = f"（{reason}）" if reason else ""
    return {
        "为什么值得关注": f"该工作属于'{category}'方向，具备行业跟踪价值。",
        "问题与背景": short_abstract or f"{no_data_msg}{reason_text}",
        "核心方法与创新": "正文抓取受限，方法细节暂无法完整还原，建议后续二次精读原文。",
        "关键结论": "正文抓取受限，暂无法给出高置信结论。",
        "增量价值与影响": f"该论文归属'{category}'方向，具备跟踪价值；当前仅可做低置信观察。",
        "局限与开放问题": "当前最大不确定性来自正文不可得或信息不足，结论可能与原文存在偏差。",
    }


def ensure_structured_analysis_content(parsed: Dict[str, str], paper: Paper, category: str) -> Dict[str, str]:
    fallback = fallback_structured_analysis(paper, category, reason="结构化字段缺失")
    out: Dict[str, str] = {}
    for k, fv in fallback.items():
        value = sanitize_text((parsed.get(k) or "").strip(), max_len=1800)
        out[k] = value if value else fv
    return out


def build_overview_lines(items: List[AnalyzedPaper]) -> List[str]:
    if not items:
        return [
            "本周总篇数：0",
            "重要性 Top 3：无",
            "本周趋势：无",
            "总体判断：本周未检索到符合条件的论文。",
        ]

    top3 = sorted(items, key=lambda x: ranking_score(x.paper), reverse=True)[:3]
    trend_pool = " ".join([normalize(f"{x.paper.title} {x.paper.abstract}") for x in items])

    trend_lines: List[str] = []
    if any(k in trend_pool for k in ["world model", "world simulator", "digital twin", "世界模型"]):
        trend_lines.append("世界模型与仿真能力继续朝可训练、可评测方向收敛")
    if any(k in trend_pool for k in ["pipeline", "etl", "ingestion", "lakehouse", "数据管道", "湖仓"]):
        trend_lines.append("数据基础设施关注管道效率、治理与可运营性")
    if any(k in trend_pool for k in ["benchmark", "evaluation", "评测", "部署", "latency"]):
        trend_lines.append("评测与部署指标被更频繁地前置到研究叙述")
    if not trend_lines:
        trend_lines = ["本周样本较少，趋势信号有限"]

    return [
        f"本周总篇数：{len(items)}",
        "重要性 Top 3：" + "；".join([f"{i+1}.{x.paper.title}" for i, x in enumerate(top3)]),
        "本周趋势：" + "；".join(trend_lines[:3]),
    ]


def to_html(report_text: str) -> str:
    lines = [ln.strip() for ln in report_text.splitlines() if ln.strip()]

    def pretty_text(val: str) -> str:
        raw = (val or "").strip()
        if not raw:
            return ""
        pieces = [x.strip() for x in re.split(r"(?<=[。！？；])", raw) if x.strip()]
        if not pieces:
            pieces = [raw]
        items = "".join([f"<li style='margin:0 0 8px 0'>{html.escape(x)}</li>" for x in pieces])
        return f"<ul style='margin:0;padding-left:22px'>{items}</ul>"

    def compact_author_line(author_raw: str) -> str:
        txt = (author_raw or "").strip()
        if not txt:
            return "未披露"
        matches = re.findall(r"([^,，（(]+)\s*[（(]([^）)]+)[）)]", txt)
        if len(matches) >= 2:
            insts = {inst.strip() for _, inst in matches if inst.strip()}
            if len(insts) == 1:
                names = [name.strip() for name, _ in matches if name.strip()]
                if names:
                    name_text = "、".join(names[:8])
                    if len(names) > 8:
                        name_text += "等"
                    inst = next(iter(insts))
                    return f"{name_text}（全部来自{inst}）"
        return txt

    title = "AI Insight - Weekly Intelligence Digest"
    subtitle = "Stay Hungry, Stay Foolish!"
    period = ""
    overview = {
        "本周总篇数": "0",
        "重要性 Top 3": "无",
        "本周趋势": "无",
    }

    papers: List[Dict[str, str]] = []
    current: Dict[str, str] = {}

    for ln in lines:
        if ln.startswith("筛选时间（北京时间）："):
            period = ln.split("：", 1)[1].strip()
        elif ln.startswith("今日总篇数："):
            overview["本周总篇数"] = ln.split("：", 1)[1].strip()
        elif ln.startswith("本周总篇数："):
            overview["本周总篇数"] = ln.split("：", 1)[1].strip()
        elif ln.startswith("重要性 Top 3："):
            overview["重要性 Top 3"] = ln.split("：", 1)[1].strip()
        elif ln.startswith("Top 3（按GitHub/X/Reddit综合重要性评分）："):
            overview["重要性 Top 3"] = ln.split("：", 1)[1].strip()
        elif ln.startswith("本周趋势："):
            overview["本周趋势"] = ln.split("：", 1)[1].strip()
        elif ln.startswith("当日趋势："):
            overview["本周趋势"] = ln.split("：", 1)[1].strip()
        elif ln.startswith("标题："):
            if current:
                papers.append(current)
            current = {
                "title": ln.split("：", 1)[1].strip(),
                "published": "",
                "author": "",
                "link": "",
                "attention": "",
                "problem": "",
                "method": "",
                "conclusion": "",
                "value": "",
                "risk": "",
            }
        elif re.match(r"^论文\d+：", ln):
            if current:
                papers.append(current)
            current = {
                "title": ln.split("：", 1)[1].strip(),
                "published": "",
                "author": "",
                "link": "",
                "attention": "",
                "problem": "",
                "method": "",
                "conclusion": "",
                "value": "",
                "risk": "",
            }
        elif current and ln.startswith("发布时间："):
            current["published"] = ln.split("：", 1)[1].strip()
        elif current and ln.startswith("作者："):
            current["author"] = ln.split("：", 1)[1].strip()
        elif current and ln.startswith("链接："):
            current["link"] = ln.split("：", 1)[1].strip()
        elif current and ln.startswith("为什么值得关注："):
            current["attention"] = ln.split("：", 1)[1].strip()
        elif current and ln.startswith("问题与背景："):
            current["problem"] = ln.split("：", 1)[1].strip()
        elif current and ln.startswith("核心方法与创新："):
            current["method"] = ln.split("：", 1)[1].strip()
        elif current and ln.startswith("关键结论："):
            current["conclusion"] = ln.split("：", 1)[1].strip()
        elif current and ln.startswith("增量价值与影响："):
            current["value"] = ln.split("：", 1)[1].strip()
        elif current and ln.startswith("局限与开放问题："):
            current["risk"] = ln.split("：", 1)[1].strip()

    if current:
        papers.append(current)

    top3 = papers[:3]
    detail_cards = []
    for idx, p in enumerate(top3, start=1):
        link = p.get("link", "")
        link_html = f"<a href='{html.escape(link)}' style='color:#2563EB;text-decoration:none'>论文链接</a>" if link else "未披露"
        detail_cards.append(
            f"""
            <tr><td style='padding:0 0 16px 0'>
              <table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='background:#FFFFFF;border:1px solid #E5E7EB;border-radius:16px;box-shadow:0 4px 16px rgba(15,23,42,0.04)'>
                <tr><td style='padding:24px'>
                  <div style='font-size:19px;line-height:1.45;font-weight:700;color:#111827;margin-bottom:8px'>{html.escape(p.get('title',''))}</div>
                  <div style='font-size:13px;color:#6B7280;line-height:1.7;margin-bottom:12px'>作者：{html.escape(compact_author_line(p.get('author','')))} ｜ {link_html}</div>

                  <div style='background:#EFF6FF;border:1px solid #BFDBFE;border-radius:12px;padding:12px 14px;margin-bottom:10px'>
                    <div style='font-size:14px;font-weight:700;color:#1D4ED8;margin-bottom:6px'>为什么值得关注</div>
                    <div style='font-size:15px;line-height:1.7;color:#111827'>{pretty_text(p.get('attention',''))}</div>
                  </div>
                  <div style='background:#F8FAFC;border:1px solid #E5E7EB;border-radius:12px;padding:12px 14px;margin-bottom:10px'>
                    <div style='font-size:14px;font-weight:700;color:#2563EB;margin-bottom:6px'>问题与背景</div>
                    <div style='font-size:15px;line-height:1.7;color:#111827'>{pretty_text(p.get('problem',''))}</div>
                  </div>
                  <div style='background:#F8FAFC;border:1px solid #E5E7EB;border-radius:12px;padding:12px 14px;margin-bottom:10px'>
                    <div style='font-size:14px;font-weight:700;color:#2563EB;margin-bottom:6px'>核心方法与创新</div>
                    <div style='font-size:15px;line-height:1.7;color:#111827'>{pretty_text(p.get('method',''))}</div>
                  </div>
                  <div style='background:#F8FAFC;border:1px solid #E5E7EB;border-radius:12px;padding:12px 14px;margin-bottom:10px'>
                    <div style='font-size:14px;font-weight:700;color:#2563EB;margin-bottom:6px'>关键结论</div>
                    <div style='font-size:15px;line-height:1.7;color:#111827'>{pretty_text(p.get('conclusion',''))}</div>
                  </div>
                  <div style='background:#F8FAFC;border:1px solid #E5E7EB;border-radius:12px;padding:12px 14px;margin-bottom:10px'>
                    <div style='font-size:14px;font-weight:700;color:#2563EB;margin-bottom:6px'>增量价值与影响</div>
                    <div style='font-size:15px;line-height:1.7;color:#111827'>{pretty_text(p.get('value',''))}</div>
                  </div>
                  <div style='background:#F8FAFC;border:1px solid #E5E7EB;border-radius:12px;padding:12px 14px'>
                    <div style='font-size:14px;font-weight:700;color:#2563EB;margin-bottom:6px'>局限与开放问题</div>
                    <div style='font-size:15px;line-height:1.7;color:#111827'>{pretty_text(p.get('risk',''))}</div>
                  </div>
                </td></tr>
              </table>
            </td></tr>
            """
        )

    overview_row = f"""
      <tr><td style='padding:0 0 8px 0;font-size:15px;line-height:1.7;color:#111827'><span style='font-weight:600'>筛选时间：</span>{html.escape(period or '未披露')}</td></tr>
      <tr><td style='padding:0 0 8px 0;font-size:15px;line-height:1.7;color:#111827'><span style='font-weight:600'>重要性 Top 3：</span>{html.escape(overview['重要性 Top 3'])}</td></tr>
      <tr><td style='padding:0;font-size:15px;line-height:1.7;color:#111827'><span style='font-weight:600'>本周趋势：</span>{html.escape(overview['本周趋势'])}</td></tr>
    """

    return f"""
<html>
  <body style='margin:0;padding:0;background:#F5F7FA;font-family:Inter,SF Pro Display,PingFang SC,Microsoft YaHei,Arial,sans-serif;color:#111827'>
    <table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='background:#F5F7FA'>
      <tr><td align='center' style='padding:24px 12px'>
        <table role='presentation' width='720' cellspacing='0' cellpadding='0' style='width:720px;max-width:720px;background:#FFFFFF;border:1px solid #E5E7EB;border-radius:16px;box-shadow:0 8px 28px rgba(15,23,42,0.06)'>
          <tr><td style='padding:32px'>

            <table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='background:#F8FAFC;border:1px solid #E5E7EB;border-radius:16px'>
              <tr><td style='padding:24px'>
                <div style='font-size:34px;line-height:1.2;font-weight:700;color:#111827;margin-bottom:8px'>{html.escape(title)}</div>
                <div style='font-size:16px;line-height:1.7;font-weight:500;color:#4B5563'>{html.escape(subtitle)}</div>
              </td></tr>
            </table>

            <div style='height:32px'></div>
            <div style='font-size:12px;font-weight:600;color:#2563EB;letter-spacing:.06em;text-transform:uppercase;margin-bottom:6px;text-align:center'>Overview</div><div style='font-size:24px;font-weight:700;color:#111827;margin-bottom:12px;text-align:center'>Weekly Overview</div>
            <table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='background:#FFFFFF;border:1px solid #E5E7EB;border-radius:16px'>
              <tr><td style='padding:20px 22px'>
                <table role='presentation' width='100%' cellspacing='0' cellpadding='0'>{overview_row}</table>
              </td></tr>
            </table>

            <div style='height:32px'></div>
            <div style='font-size:12px;font-weight:600;color:#2563EB;letter-spacing:.06em;text-transform:uppercase;margin-bottom:6px;text-align:center'>Paper Deep Dive</div><div style='font-size:24px;font-weight:700;color:#111827;margin-bottom:12px;text-align:center'>论文详解</div>
            <table role='presentation' width='100%' cellspacing='0' cellpadding='0'>{''.join(detail_cards) if detail_cards else "<tr><td style='font-size:16px;line-height:1.7;color:#4B5563'>本期暂无可解析论文。</td></tr>"}</table>

            <div style='height:32px'></div>
            <!-- WEEKLY_SIGNALS_SLOT -->

            <div style='height:18px'></div>
            <div style='font-size:12px;line-height:1.6;color:#6B7280'>AI Insight internal use only · auto-generated weekly intelligence digest</div>

          </td></tr>
        </table>
      </td></tr>
    </table>
  </body>
</html>
""".strip()


def _export_paper_quality_checkpoint(papers: List[Paper]) -> None:
    """Export all fetched papers to Excel with score breakdowns, sorted by final score.

    This serves as a quality checkpoint: every paper within the time window
    is listed with its topical / social / quality scores so reviewers can
    verify the ranking logic and spot any anomalies.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[QC] openpyxl not installed, skipping paper quality checkpoint export")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Paper Quality Checkpoint"

    headers = [
        "排名", "标题", "作者", "原文链接",
        "主题相关分", "社交热度分", "质量评估分", "综合排序分",
    ]
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(bottom=Side(style="thin", color="D1D5DB"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    scored_papers = []
    for p in papers:
        t_score = topical_score(p.title, p.abstract)
        s_score = float(getattr(p, "_social_score", 0.0))
        if s_score <= 0:
            try:
                s_score, details = compute_social_discussion_score(p)
                setattr(p, "_social_score", float(s_score))
                setattr(p, "_social_details", details)
            except Exception:
                s_score = 0.0
        q_score = 0
        try:
            q = compute_early_quality_score(p, classify_paper(p), (p.abstract or "").strip())
            q_score = int(((q.get("scores") or {}).get("total_score") or 0))
        except Exception:
            q_score = 0
        r_score = ranking_score(p)
        scored_papers.append((p, t_score, s_score, q_score, r_score))

    scored_papers.sort(key=lambda x: x[4], reverse=True)

    for idx, (p, t_sc, s_sc, q_sc, r_sc) in enumerate(scored_papers, 2):
        ws.cell(row=idx, column=1, value=idx - 1)
        ws.cell(row=idx, column=2, value=p.title or "")
        ws.cell(row=idx, column=3, value=", ".join(p.authors[:5]) if p.authors else "")
        ws.cell(row=idx, column=4, value=p.url or "")
        ws.cell(row=idx, column=5, value=t_sc)
        ws.cell(row=idx, column=6, value=round(s_sc, 1))
        ws.cell(row=idx, column=7, value=q_sc)
        ws.cell(row=idx, column=8, value=round(r_sc, 2))
        for col in range(1, len(headers) + 1):
            ws.cell(row=idx, column=col).border = thin_border
            ws.cell(row=idx, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.auto_filter.ref = ws.dimensions

    dest = PAPERS_DIR / "paper_quality_checkpoint.xlsx"
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    print(f"[QC] paper quality checkpoint exported: {dest} ({len(scored_papers)} papers)")


def build_daily_digest(client: OpenAI) -> Tuple[str, str]:
    papers, _ = collect_recent_papers()

    if not papers:
        start, end = target_beijing_date_window()
        text = (
            "World Engine 与 Data Infra 论文周报\n"
            f"筛选时间（北京时间）：{start.strftime('%Y-%m-%d')} 至 {end.strftime('%Y-%m-%d')}\n"
            "本周总篇数：0\n"
            "重要性 Top 3：无\n"
            "本周趋势：无"
        )
        cleaned = clean_symbols(text)
        return cleaned, to_html(cleaned)

    # ── Export candidate papers to Excel with full score breakdowns first ──
    _export_paper_quality_checkpoint(papers)

    # Top-3 MUST follow the final composite score used in Excel ("综合排序分")
    # i.e. ranking_score descending on the same candidate paper set.
    candidate_pool = sorted(papers, key=ranking_score, reverse=True)
    selected = candidate_pool[:3]
    for p in selected:
        # Ensure social context exists for the "为什么值得关注" section.
        if not getattr(p, "_social_details", None):
            try:
                social_score, details = compute_social_discussion_score(p)
                setattr(p, "_social_score", social_score)
                setattr(p, "_social_details", details)
            except Exception:
                pass

    analyzed: List[AnalyzedPaper] = []
    parsed_map: Dict[str, Dict[str, str]] = {}
    score_detail_map: Dict[str, Dict[str, object]] = {}
    skipped_no_fulltext = 0
    abstract_fallback_used = 0
    # Download PDFs for top 3 papers into papers/ directory
    pdf_dir = PAPERS_DIR
    pdf_dir.mkdir(parents=True, exist_ok=True)
    print(f"[INFO] PDF output directory: {pdf_dir.resolve()}")

    for paper_idx, paper in enumerate(selected):
        category = classify_paper(paper)
        # Delay between downloads to avoid arXiv rate-limiting
        if paper_idx > 0:
            import time; time.sleep(3)
        # Primary: download PDF and extract full text
        print(f"[INFO] Processing paper {paper_idx + 1}/{len(selected)}: {(paper.title or '')[:60]}")
        pdf_text, pdf_path = fetch_fulltext_via_pdf(paper, dest_dir=pdf_dir)
        if len(pdf_text.strip()) >= 2000:
            fulltext_context = pdf_text
            print(f"[DEEP] full PDF text for: {paper.title[:60]} ({len(pdf_text)} chars)")
        else:
            # Fallback: HTML scraping then abstract
            fulltext_context = fetch_fulltext_context(paper)
        analysis_context = fulltext_context
        if not has_readable_fulltext(fulltext_context):
            skipped_no_fulltext += 1
            analysis_context = (paper.abstract or "").strip()
            abstract_fallback_used += 1
        score_json = compute_early_quality_score(paper, category, analysis_context)
        early_score = int(((score_json.get("scores") or {}).get("total_score") or 0))
        if analysis_context:
            try:
                raw = analyze_paper(client, paper, category, analysis_context)
                parsed = ensure_structured_analysis_content(parse_structured_analysis(raw), paper, category)
            except Exception as exc:
                parsed = fallback_structured_analysis(paper, category, reason=f"LLM解析失败：{exc}")
        else:
            parsed = fallback_structured_analysis(paper, category, reason="正文与摘要均不足")
        analyzed.append(AnalyzedPaper(paper=paper, category=category, analysis_lines=[], early_score=early_score, discussion_score=float(getattr(paper, "_social_score", 0.0))))
        parsed_map[paper.title] = parsed
        score_detail_map[paper.title] = score_json

    if not analyzed:
        start, end = target_beijing_date_window()
        text = (
            "World Engine 与 Data Infra 论文周报\n"
            f"筛选时间（北京时间）：{start.strftime('%Y-%m-%d')} 至 {end.strftime('%Y-%m-%d')}\n"
            "本周总篇数：0\n"
            "重要性 Top 3：无\n"
            f"本周趋势：候选论文正文抓取不足（{skipped_no_fulltext}篇），信息不足未生成解读"
        )
        cleaned = clean_symbols(text)
        return cleaned, to_html(cleaned)


    analyzed.sort(key=lambda x: ranking_score(x.paper), reverse=True)
    rank_map = {it.paper.title: i + 1 for i, it in enumerate(analyzed)}

    start, end = target_beijing_date_window()
    blocks: List[str] = [
        "World Engine 与 Data Infra 论文周报",
        f"筛选时间（北京时间）：{start.strftime('%Y-%m-%d')} 至 {end.strftime('%Y-%m-%d')}",
    ]
    blocks.extend(build_overview_lines(analyzed))

    n = 1
    for item in analyzed:
        blocks.extend(render_paper_block(n, item, parsed_map[item.paper.title], rank_map.get(item.paper.title, n)))
        n += 1

    text = clean_symbols("\n".join(blocks))
    return text, to_html(text)


def build_official_monitor_section() -> Tuple[str, str]:
    if os.environ.get("OFFICIAL_MONITOR_ENABLED", "1").strip() in {"0", "false", "False"}:
        return "", ""

    try:
        try:
            from agent.official_monitor.pipeline import run_pipeline
            from agent.official_monitor.render import render_markdown, render_html_fragment
        except ModuleNotFoundError:
            from official_monitor.pipeline import run_pipeline
            from official_monitor.render import render_markdown, render_html_fragment

        lookback = int(os.environ.get("OFFICIAL_MONITOR_LOOKBACK_DAYS", "7"))
        max_per_source = int(os.environ.get("OFFICIAL_MONITOR_MAX_PER_SOURCE", "20"))
        summary, deduped_articles, clusters, kept_articles, reflection_result = run_pipeline(lookback_days=lookback, max_articles_per_source=max_per_source)

        # Store monitor metrics for run_history (consumed by run_once)
        build_official_monitor_section._last_metrics = {
            "signal_articles_fetched": summary.fetched_articles,
            "signal_articles_deduped": summary.deduped_articles,
            "signal_articles_kept": summary.kept_articles,
            "signal_clusters": summary.topic_clusters,
            "signal_drop_reasons": summary.drop_reasons,
            "reflection": reflection_result,
        }

        if not clusters:
            text = "本周 AI 官方信号图谱\n本周无核心异动"
            html_block = (
                "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='margin-top:30px'>"
                "<tr><td style='background:#FFFFFF;border:1px solid #E5E7EB;border-radius:16px;padding:24px'>"
                "<div style='font-size:24px;line-height:1.25;font-weight:700;color:#111827;margin-bottom:8px'>本周 AI 官方信号图谱</div>"
                "<div style='font-size:14px;line-height:1.6;color:#4B5563;margin-bottom:12px'>来自 AI 大厂与投资机构官网的主题归纳</div>"
                "<div style='font-size:16px;line-height:1.7;color:#111827'>本周无核心异动</div>"
                "</td></tr></table>"
            )
            return text, html_block

        md = render_markdown(summary, clusters)
        html_fragment = render_html_fragment(summary, clusters)
        text = "本周 AI 官方信号图谱\n" + md
        return text, html_fragment
    except Exception as exc:
        warn = f"本周 AI 官方信号图谱\n该板块本次生成失败：{exc}"
        warn_html = (
            "<table role='presentation' width='100%' cellspacing='0' cellpadding='0' style='margin-top:30px'>"
            "<tr><td style='background:#FFFFFF;border:1px solid #E5E7EB;border-radius:16px;padding:24px'>"
            "<div style='font-size:24px;line-height:1.25;font-weight:700;color:#111827;margin-bottom:8px'>本周 AI 官方信号图谱</div>"
            "<div style='font-size:14px;line-height:1.6;color:#4B5563;margin-bottom:12px'>来自 AI 大厂与投资机构官网的主题归纳</div>"
            f"<div style='font-size:16px;line-height:1.7;color:#B91C1C'>该板块本次生成失败：{html.escape(str(exc))}</div>"
            "</td></tr></table>"
        )
        return warn, warn_html



def send_email(subject: str, text_body: str, html_body: str) -> None:
    raw_to_email = os.environ["REPORT_EMAIL_TO"]
    to_emails = [
        addr.strip()
        for addr in re.split(r"[;,]", raw_to_email)
        if addr and addr.strip()
    ]
    if not to_emails:
        raise ValueError("REPORT_EMAIL_TO is empty after parsing")

    from_email = os.environ.get("REPORT_EMAIL_FROM", to_emails[0])
    smtp_host = os.environ.get("SMTP_HOST", "smtp.163.com")
    smtp_port = int(os.environ.get("SMTP_PORT", "465"))
    smtp_user = os.environ.get("SMTP_USER", from_email)
    smtp_pass = os.environ.get("SMTP_PASS", "")

    msg = MIMEMultipart("alternative")
    msg["From"] = from_email
    msg["To"] = ", ".join(to_emails)
    msg["Subject"] = subject
    msg.attach(MIMEText(text_body, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    with smtplib.SMTP_SSL(smtp_host, smtp_port) as server:
        if smtp_user and smtp_pass:
            server.login(smtp_user, smtp_pass)
        server.sendmail(from_email, to_emails, msg.as_string())


def run_once() -> None:
    client = OpenAI(
        api_key=os.environ["GOOGLE_API_KEY"],
        base_url=GEMINI_BASE_URL,
    )
    text_digest, html_digest = build_daily_digest(client)
    official_text, official_html = build_official_monitor_section()
    if official_text:
        text_digest = text_digest + "\n\n" + official_text
    if official_html:
        if "<!-- WEEKLY_SIGNALS_SLOT -->" in html_digest:
            html_digest = html_digest.replace("<!-- WEEKLY_SIGNALS_SLOT -->", official_html, 1)
        elif "</body>" in html_digest:
            html_digest = html_digest.replace("</body>", official_html + "</body>", 1)
        else:
            html_digest += official_html
    else:
        html_digest = html_digest.replace("<!-- WEEKLY_SIGNALS_SLOT -->", "", 1)
    date_str = now_beijing().strftime("%Y-%m-%d")
    send_email(
        subject=f"[{date_str}] World Engine 与 Data Infra 每周论文简报",
        text_body=text_digest,
        html_body=html_digest,
    )
    print("[OK] weekly digest sent")

    # ── Record run metrics to run_history.jsonl ──
    try:
        from run_history import record_run
        monitor_metrics = getattr(build_official_monitor_section, "_last_metrics", {})
        record_run(
            signal_articles_fetched=monitor_metrics.get("signal_articles_fetched", 0),
            signal_articles_deduped=monitor_metrics.get("signal_articles_deduped", 0),
            signal_articles_kept=monitor_metrics.get("signal_articles_kept", 0),
            signal_clusters=monitor_metrics.get("signal_clusters", 0),
            signal_drop_reasons=monitor_metrics.get("signal_drop_reasons"),
            reflection=monitor_metrics.get("reflection"),
        )
    except Exception as exc:
        print(f"[WARN] failed to record run history: {exc}")


def run_scheduler() -> None:
    report_time = os.environ.get("REPORT_TIME", "10:00")
    hour, minute = [int(x) for x in report_time.split(":", 1)]
    timezone = os.environ.get("TZ", "Asia/Shanghai")

    scheduler = BlockingScheduler(timezone=timezone)
    scheduler.add_job(run_once, "cron", day_of_week="mon", hour=hour, minute=minute)
    print(f"[INFO] scheduler started, every Monday at {report_time} ({timezone})")
    scheduler.start()


if __name__ == "__main__":
    if os.environ.get("AGENT_MODE", "once").strip().lower() == "schedule":
        run_scheduler()
    else:
        run_once()
